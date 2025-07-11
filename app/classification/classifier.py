import spacy
from datetime import datetime
from classification.departments import DEPARTMENTS, HIGH_PRIORITY_KEYWORDS, get_department_id_by_name
import pytesseract
from PIL import Image
import io
import pdfplumber
from docx import Document as DocxDocument
import openpyxl

nlp = spacy.load("en_core_web_sm")

# Configure Tesseract for better OCR accuracy
pytesseract.pytesseract.tesseract_cmd = '/usr/bin/tesseract'

def is_high_priority(text: str) -> bool:
    """Checks for urgency based on priority keywords"""
    return any(keyword.lower() in text.lower() for keyword in HIGH_PRIORITY_KEYWORDS)

def extract_text_from_image(image_bytes: bytes) -> str:
    """Extracts text via OCR from image bytes with improved accuracy"""
    try:
        image = Image.open(io.BytesIO(image_bytes))
        
        # Improve image quality for better OCR
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        # Resize image if too small
        width, height = image.size
        if width < 300 or height < 300:
            scale_factor = max(300/width, 300/height)
            new_size = (int(width * scale_factor), int(height * scale_factor))
            image = image.resize(new_size, Image.Resampling.LANCZOS)
        
        # Use custom OCR config for better accuracy
        custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz.,!?:;()-[]{}"\' '
        text = pytesseract.image_to_string(image, config=custom_config)
        return text.strip()
    except Exception as e:
        print(f"OCR error: {e}")
        return ""

def extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extract text from PDF file"""
    try:
        text = ""
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text.strip()
    except Exception as e:
        print(f"PDF extraction error: {e}")
        return ""

def extract_text_from_docx(file_bytes: bytes) -> str:
    """Extract text from DOCX file"""
    try:
        doc = DocxDocument(io.BytesIO(file_bytes))
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    except Exception as e:
        print(f"DOCX extraction error: {e}")
        return ""

def extract_text_from_xlsx(file_bytes: bytes) -> str:
    """Extract text from Excel file"""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    text += row_text + "\n"
        return text.strip()
    except Exception as e:
        print(f"Excel extraction error: {e}")
        return ""

def extract_text_from_file(file_bytes: bytes, filename: str) -> str:
    """Extract text from various file formats"""
    file_extension = filename.lower().split('.')[-1] if '.' in filename else ''
    
    if file_extension in ['jpg', 'jpeg', 'png', 'bmp', 'tiff', 'gif']:
        return extract_text_from_image(file_bytes)
    elif file_extension == 'pdf':
        return extract_text_from_pdf(file_bytes)
    elif file_extension in ['docx', 'doc']:
        return extract_text_from_docx(file_bytes)
    elif file_extension in ['xlsx', 'xls']:
        return extract_text_from_xlsx(file_bytes)
    elif file_extension in ['txt', 'csv']:
        try:
            return file_bytes.decode('utf-8', errors='ignore')
        except:
            return ""
    else:
        # Try to decode as text first
        try:
            return file_bytes.decode('utf-8', errors='ignore')
        except:
            return ""

def classify_document(text: str, filename: str = "") -> dict:
    """Classifies document based on keyword counts across departments"""
    if not text.strip():
        return {
            "departments": ["general"],
            "department_ids": [9],  # General department ID
            "priority": "Normal",
            "date": datetime.today().date()
        }
    
    doc = nlp(text.lower())
    keyword_counts = {}

    for dept_id, dept_info in DEPARTMENTS.items():
        dept_name = dept_info["name"]
        keywords = dept_info["keywords"]
        count = sum(text.lower().count(keyword.lower()) for keyword in keywords)
        if count > 0:
            keyword_counts[dept_name] = {"count": count, "id": dept_id}

    if keyword_counts:
        sorted_departments = sorted(keyword_counts.items(), key=lambda item: item[1]["count"], reverse=True)
        top_departments = [dept for dept, _ in sorted_departments[:2]]
        top_department_ids = [keyword_counts[dept]["id"] for dept in top_departments]
    else:
        top_departments = ["general"]
        top_department_ids = [9]  # General department ID

    priority = "High" if is_high_priority(text) else "Normal"
    extracted_date = datetime.today().date()

    return {
        "departments": top_departments,
        "department_ids": top_department_ids,
        "priority": priority,
        "date": extracted_date
    }
